from openpyxl import Workbook, load_workbook
import os


def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.create_sheet('Sheet1')
    # sheet = workbook.active

    # Add headers (heading of table)
    sheet.append(["Name", "Age"])

    # Add Data
    sheet.append(['John doe', '30'])
    sheet.append(['Jane smith', '25'])

    # save
    workbook.save(filename)
    workbook.close()
    print("Data write and File Closed Successfully")


def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook['Sheet1']
    for row in sheet.iter_rows(2, values_only=True):
        print(f"Name:{row[0]}, Age: {row[1]}")
    workbook.close()
    print("Workbook Closed")


def delete_excel(filename):
    workbook = load_workbook(filename)
    if workbook.active:
        print('Active Workbook ')
        workbook.close()

    try:
        os.remove(filename)
        print(f"{filename} deleted Successfully")
    except FileNotFoundError:
        print(f"{filename} does not Exist")
    except PermissionError:
        print(f" Permission error : {filename} is currently in use")


if __name__ == "__main__":
    filename = "data.xlsx"
    write_excel(filename)
    print("Data read from xlsx file : ")
    read_excel(filename)
